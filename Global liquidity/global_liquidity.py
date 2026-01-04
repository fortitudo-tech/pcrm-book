# ======================================================================
# DIY Global Liquidity Index (GLI) Script
# ======================================================================
# Requirements (run once in your environment, e.g. in a notebook cell):
#   pip install pandas pandas_datareader yfinance matplotlib openpyxl xlsxwriter
# ======================================================================

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
from pandas_datareader import data as web
import yfinance as yf
from datetime import datetime

# Make plots a bit nicer
plt.rcParams["figure.figsize"] = (12, 6)
plt.rcParams["axes.grid"] = True

# ----------------------------------------------------------------------
# 1. Parameters
# ----------------------------------------------------------------------
START_DATE = "2010-01-01"
EXCEL_OUTPUT = "global_liquidity_DIY_v2.xlsx"

# ----------------------------------------------------------------------
# 2. Download central-bank balance sheet data from FRED
# ----------------------------------------------------------------------
# FRED series IDs:
#   Fed: WALCL          (Total assets)
#   ECB: ECBASSETSW     (Assets of ECB/Eurosystem, weekly - updated series)
#   BoJ: JPNASSETS      (Bank of Japan total assets)
#   BoE: BOEASSETS      (Bank of England total assets - may not be available)

cb_series = {
    "Fed": "WALCL",
    "ECB": "ECBASSETSW",
    "BOJ": "JPNASSETS"
    # BOE removed - no current weekly series available on FRED
}

print("Downloading central bank balance sheets from FRED...")
cb_list = []
cb_names = []

for name, code in cb_series.items():
    try:
        series = web.DataReader(code, "fred", START_DATE)
        cb_list.append(series)
        cb_names.append(name)
        print(f"  - {name}: {code} OK")
    except Exception as e:
        print(f"  - {name}: FAILED ({code}) -> {e}")

# Concatenate all series with outer join, then forward fill missing values
cb = pd.concat(cb_list, axis=1, keys=cb_names)
# Flatten the multi-index columns to single level
cb.columns = cb.columns.get_level_values(0)
print(f"Central bank data before ffill: {cb.shape}")
cb = cb.ffill()  # Forward fill missing values (each CB reports at different frequencies)
cb = cb.dropna()  # Drop rows where data doesn't exist yet
print("Central bank data shape:", cb.shape)

# ----------------------------------------------------------------------
# 3. FX series to convert ECB/BoJ to USD
# ----------------------------------------------------------------------
# FRED FX series:
#   DEXUSEU: USD per EUR
#   DEXJPUS: JPY per USD

print("Downloading FX data from FRED...")
eurusd = web.DataReader("DEXUSEU", "fred", START_DATE)   # USD per EUR
jpyusd = web.DataReader("DEXJPUS", "fred", START_DATE)   # JPY per USD

fx = pd.concat([eurusd, jpyusd], axis=1)
fx.columns = ["EURUSD", "JPYUSD"]
fx = fx.ffill()

print("FX data shape:", fx.shape)

# Align CB and FX dates
data = cb.join(fx, how="inner")
print("Merged CB + FX shape:", data.shape)
print("Column names:", list(data.columns))

# ----------------------------------------------------------------------
# 4. Convert CB balance sheets to USD and compute total CB liquidity
# ----------------------------------------------------------------------
# Assumes ECB/BoE/BoJ local series are in their domestic currencies
#   ECB: EUR, BoE: GBP, BoJ: JPY

data["ECB_USD"] = data["ECB"] / data["EURUSD"]
if "BOE" in data.columns:
    data["BOE_USD"] = data["BOE"] / data["GBPUSD"]
    cb_usd_cols = ["Fed", "ECB_USD", "BOE_USD", "BOJ_USD"]
else:
    print("Warning: BOE data not available, excluding from calculations")
    cb_usd_cols = ["Fed", "ECB_USD", "BOJ_USD"]
data["BOJ_USD"] = data["BOJ"] / data["JPYUSD"]

data["CB_Total"] = data[cb_usd_cols].sum(axis=1)

print("Computed CB_Total in USD.")

# ----------------------------------------------------------------------
# 5. VIX (volatility) and DXY (broad dollar index)
# ----------------------------------------------------------------------
# FRED series:
#   VIXCLS: CBOE Volatility Index (VIX) - stock market volatility (MOVE not available)
#   DTWEXBGS: broad trade-weighted USD index

print("Downloading VIX and DXY from FRED...")
try:
    move = web.DataReader("VIXCLS", "fred", START_DATE)
    has_move = True
    print("  - VIX: OK")
except Exception as e:
    print(f"  - VIX: FAILED -> {e}")
    has_move = False

try:
    dxy = web.DataReader("DTWEXBGS", "fred", START_DATE)
    has_dxy = True
    print("  - DXY: OK")
except Exception as e:
    print(f"  - DXY: FAILED -> {e}")
    has_dxy = False

# Join available data
if has_move:
    data = data.join(move, how="inner")
if has_dxy:
    data = data.join(dxy, how="inner")

# Dynamically construct column names based on available data
additional_cols = ["ECB_USD"]
if "BOE" in cb.columns:
    additional_cols.append("BOE_USD")
additional_cols.extend(["BOJ_USD", "CB_Total"])
if has_move:
    additional_cols.append("VIX")
if has_dxy:
    additional_cols.append("DXY")
data.columns = list(cb.columns) + list(fx.columns) + additional_cols

data = data.dropna()
print("Final merged daily data shape:", data.shape)

# ----------------------------------------------------------------------
# 6. Resample to weekly (Flash-style) and build GLI_Flash + GLI_Full
# ----------------------------------------------------------------------
print("Resampling to weekly (Friday close)...")
weekly = data.resample("W-FRI").last()
weekly = weekly.dropna()
print("Weekly data shape:", weekly.shape)

def full_sample_zscore(x: pd.Series) -> pd.Series:
    """
    Calculate z-score using the FULL sample mean and std.
    This creates a stable index across the entire history.
    """
    return (x - x.mean()) / x.std()

def clip_zscore(z: pd.Series, clip: float = 3.0) -> pd.Series:
    """
    Clip extreme z-scores to prevent COVID-like events from compressing the chart.
    """
    return z.clip(lower=-clip, upper=clip)

# GLI_Flash: based on short-term MOMENTUM (13-week change)
# This captures acceleration/deceleration in liquidity, not absolute levels
cb_change_13w = weekly["CB_Total"].pct_change(13)  # 13-week momentum
gli_flash = full_sample_zscore(cb_change_13w)
gli_flash = clip_zscore(gli_flash, clip=3.0)

# Apply light smoothing (4-week MA) to reduce noise while preserving responsiveness
gli_flash = gli_flash.rolling(window=4, min_periods=1).mean()

# Subtract volatility (VIX) - high volatility reduces effective liquidity
if has_move and "VIX" in weekly.columns:
    vix_z = full_sample_zscore(weekly["VIX"])
    vix_z = clip_zscore(vix_z, clip=3.0)
    gli_flash = gli_flash - vix_z
else:
    print("Warning: VIX data not available, GLI_Flash calculated without volatility")

# Subtract dollar strength (DXY) - strong dollar tightens global liquidity
if has_dxy and "DXY" in weekly.columns:
    dxy_z = full_sample_zscore(weekly["DXY"])
    dxy_z = clip_zscore(dxy_z, clip=3.0)
    gli_flash = gli_flash - dxy_z
else:
    print("Warning: DXY data not available, GLI_Flash calculated without dollar index")

weekly["GLI_Flash"] = gli_flash

# GLI_Full: smoothed version using 20-week simple moving average
# This shows the slow-moving macro regime with better responsiveness
weekly["GLI_Full"] = weekly["GLI_Flash"].rolling(window=20, min_periods=10).mean()

print("Computed GLI_Flash and GLI_Full.")

# ----------------------------------------------------------------------
# 7. Growth metrics (YoY and 3m annualized on CB_Total)
# ----------------------------------------------------------------------
# Use smoothed values to avoid noisy denominators causing extreme spikes

# Smooth CB_Total with 4-week average for more stable growth calculations
cb_smoothed = weekly["CB_Total"].rolling(window=4, min_periods=1).mean()

# YoY growth: compare current 4-week average to 52 weeks ago
weekly["CB_YoY"] = cb_smoothed.pct_change(52)

# 3-month annualized: compare current 4-week average to 13 weeks ago, then annualize
# Clip raw growth to prevent extreme values when denominator is very small
growth_3m = cb_smoothed.pct_change(13).clip(lower=-0.2, upper=0.3)  # Clip to ±20% to 30%
# Annualize: (1 + 13-week growth)^(52/13) - 1, then clip annualized result
cb_3m_ann = ((1 + growth_3m) ** (52/13)) - 1
weekly["CB_3m_ann"] = cb_3m_ann.clip(lower=-0.6, upper=1.5)  # Clip annualized to ±60% to 150%

print("Computed CB_YoY and CB_3m_ann growth rates.")

# ----------------------------------------------------------------------
# 8. MSCI overlay using ACWI as proxy (Yahoo Finance)
# ----------------------------------------------------------------------
print("Downloading ACWI (MSCI proxy) from Yahoo Finance...")
acwi_data = yf.download("ACWI", start=START_DATE, progress=False)

# Handle both single and multi-column formats
if isinstance(acwi_data.columns, pd.MultiIndex):
    # MultiIndex columns - get the first level value
    if "Adj Close" in acwi_data.columns.get_level_values(0):
        acwi = acwi_data.xs("Adj Close", axis=1, level=0)
        if isinstance(acwi, pd.DataFrame):
            acwi = acwi.iloc[:, 0]
    else:
        acwi = acwi_data.xs("Close", axis=1, level=0)
        if isinstance(acwi, pd.DataFrame):
            acwi = acwi.iloc[:, 0]
elif "Adj Close" in acwi_data.columns:
    acwi = acwi_data["Adj Close"]
else:
    # Try Close if Adj Close not available
    acwi = acwi_data["Close"]

acwi.name = "ACWI"

weekly = weekly.join(acwi.resample("W-FRI").last())
weekly = weekly.dropna(subset=["ACWI"])
print("Weekly data with ACWI shape:", weekly.shape)

# Normalize to 0-1 using full-sample min/max (ONE TIME, not rolling)
def normalize_full_sample(x: pd.Series) -> pd.Series:
    """
    Normalize to 0-1 using the full sample min and max.
    This creates a consistent scale across the entire history.
    """
    return (x - x.min()) / (x.max() - x.min())

# Use GLI_Full (smoothed) instead of GLI_Flash for clearer overlay
overlay = pd.DataFrame({
    "GLI": normalize_full_sample(weekly["GLI_Full"]),
    "MSCI": normalize_full_sample(weekly["ACWI"])
}).dropna()

print("Overlay data shape:", overlay.shape)

# ----------------------------------------------------------------------
# 9. Plots (Three versions: Yearly, Monthly, Weekly ticks)
# ----------------------------------------------------------------------
print("Generating charts...")

import matplotlib.dates as mdates

def create_flash_vs_full_chart(tick_type='yearly', suffix=''):
    """Create Flash vs Full chart with specified tick granularity"""
    fig, ax = plt.subplots(figsize=(12, 6))

    # Filter data based on time range
    if tick_type == 'weekly':
        # Last 2 years for weekly view
        cutoff_date = pd.Timestamp.now() - pd.DateOffset(years=2)
        data_subset = weekly[weekly.index >= cutoff_date]
    elif tick_type == 'monthly':
        # Last 5-7 years for monthly view (using 6 years)
        cutoff_date = pd.Timestamp.now() - pd.DateOffset(years=6)
        data_subset = weekly[weekly.index >= cutoff_date]
    else:  # yearly - keep full history
        data_subset = weekly

    # Plot with improved visual separation: Flash = thin solid blue, Full = thick solid orange
    ax.plot(data_subset.index, data_subset["GLI_Flash"],
            label="GLI Flash", color="blue", linewidth=1, alpha=0.7)
    ax.plot(data_subset.index, data_subset["GLI_Full"],
            label="GLI Full (Smoothed)", color="orange", linewidth=2.5)

    ax.set_title(f"Global Liquidity – Flash vs Full ({tick_type.capitalize()} View)")
    ax.set_xlabel("Date")
    ax.set_ylabel("Index (z-score based)")
    ax.legend()
    ax.grid(True, alpha=0.3)

    # Set appropriate date formatter based on tick type
    if tick_type == 'yearly':
        ax.xaxis.set_major_locator(mdates.YearLocator())
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
    elif tick_type == 'monthly':
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))  # Every 3 months
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%b\n%Y'))
    else:  # weekly
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=1))  # Every month for weekly view
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%b\n%Y'))

    plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')
    plt.tight_layout()
    plt.savefig(f'gli_plot_flash_vs_full{suffix}.png', dpi=150, bbox_inches='tight')
    plt.close()

# Generate all three versions
create_flash_vs_full_chart('yearly', '_yearly')
create_flash_vs_full_chart('monthly', '_monthly')
create_flash_vs_full_chart('weekly', '_weekly')

def create_growth_chart(tick_type='yearly', suffix=''):
    """Create growth chart with specified tick granularity"""
    fig, ax1 = plt.subplots(figsize=(12, 6))

    # Filter data based on time range
    if tick_type == 'weekly':
        # Last 2 years for weekly view
        cutoff_date = pd.Timestamp.now() - pd.DateOffset(years=2)
        data_subset = weekly[weekly.index >= cutoff_date]
    elif tick_type == 'monthly':
        # Last 5-7 years for monthly view (using 6 years)
        cutoff_date = pd.Timestamp.now() - pd.DateOffset(years=6)
        data_subset = weekly[weekly.index >= cutoff_date]
    else:  # yearly - keep full history
        data_subset = weekly

    ax1.plot(data_subset.index, data_subset["CB_3m_ann"] * 100,
             label="3-Month Annualized Liquidity Growth", alpha=0.7)
    ax1.set_ylabel("3-Month Annualized (%)")
    ax1.set_xlabel("Date")
    ax1.grid(True, alpha=0.3)

    ax2 = ax1.twinx()
    ax2.plot(data_subset.index, data_subset["CB_YoY"] * 100,
             color="orange", label="Liquidity YoY Change", linewidth=2)
    ax2.set_ylabel("Year-over-Year (%)")

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper left")

    # Set appropriate date formatter based on tick type
    if tick_type == 'yearly':
        ax1.xaxis.set_major_locator(mdates.YearLocator())
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
    elif tick_type == 'monthly':
        ax1.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%b\n%Y'))
    else:  # weekly
        ax1.xaxis.set_major_locator(mdates.MonthLocator(interval=1))
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%b\n%Y'))

    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
    plt.title(f"Global Liquidity Growth ({tick_type.capitalize()} View)")
    plt.tight_layout()
    plt.savefig(f'gli_plot_growth{suffix}.png', dpi=150, bbox_inches='tight')
    plt.close()

# Generate all three versions
create_growth_chart('yearly', '_yearly')
create_growth_chart('monthly', '_monthly')
create_growth_chart('weekly', '_weekly')

def create_overlay_chart(tick_type='yearly', suffix=''):
    """Create overlay chart with specified tick granularity"""
    fig, ax = plt.subplots(figsize=(12, 6))

    # Filter data based on time range
    if tick_type == 'weekly':
        # Last 2 years for weekly view
        cutoff_date = pd.Timestamp.now() - pd.DateOffset(years=2)
        data_subset = overlay[overlay.index >= cutoff_date]
    elif tick_type == 'monthly':
        # Last 5-7 years for monthly view (using 6 years)
        cutoff_date = pd.Timestamp.now() - pd.DateOffset(years=6)
        data_subset = overlay[overlay.index >= cutoff_date]
    else:  # yearly - keep full history for macro comparison
        data_subset = overlay

    ax.plot(data_subset.index, data_subset["MSCI"], label="MSCI ACWI (Global Equities)", color="black", linewidth=2)
    ax.plot(data_subset.index, data_subset["GLI"], label="Global Liquidity Index (Full)", color="orange", linewidth=2)
    ax.set_title(f"Global Liquidity vs. Equities ({tick_type.capitalize()} View)")
    ax.set_xlabel("Date")
    ax.set_ylabel("Normalized Index (0-1)")
    ax.legend(loc="upper left")
    ax.grid(True, alpha=0.3)

    # Set appropriate date formatter based on tick type
    if tick_type == 'yearly':
        ax.xaxis.set_major_locator(mdates.YearLocator())
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
    elif tick_type == 'monthly':
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%b\n%Y'))
    else:  # weekly
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=1))
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%b\n%Y'))

    plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')
    plt.tight_layout()
    plt.savefig(f'gli_plot_msci_overlay{suffix}.png', dpi=150, bbox_inches='tight')
    plt.close()

# Generate all three versions
create_overlay_chart('yearly', '_yearly')
create_overlay_chart('monthly', '_monthly')
create_overlay_chart('weekly', '_weekly')

print("Generated 9 charts total (3 chart types × 3 time granularities)")

# ----------------------------------------------------------------------
# 10. Export to Excel with conditional formatting and helper columns
# ----------------------------------------------------------------------
print(f"Writing data to Excel: {EXCEL_OUTPUT}")

from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

# Add helper columns to weekly data
weekly_export = weekly.copy()

# Helper Column 1: Liquidity % Change (week-over-week change in CB_Total)
weekly_export["Liquidity_Pct_Change"] = weekly_export["CB_Total"].pct_change() * 100

# Helper Column 2: Trend Signal
def get_trend_signal(pct_change):
    if pd.isna(pct_change):
        return "Flat"
    elif pct_change > 1.0:
        return "Rising"
    elif pct_change < -1.0:
        return "Falling"
    else:
        return "Flat"

weekly_export["Trend_Signal"] = weekly_export["Liquidity_Pct_Change"].apply(get_trend_signal)

# Write to Excel
with pd.ExcelWriter(EXCEL_OUTPUT, engine="openpyxl") as writer:
    weekly_export.to_excel(writer, sheet_name="Weekly_Liquidity")
    overlay.to_excel(writer, sheet_name="MSCI_Overlay")

    # Get the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets["Weekly_Liquidity"]

    # Apply 3-color gradient to data columns (skip Date column which is index, now in column A)
    # Find the last row and column with data
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    # Get column letters for all numeric data columns (B onwards, excluding helper columns at end)
    # Columns are: A=Date, B=Fed, C=ECB, D=BOJ, E=EURUSD, F=JPYUSD, G=ECB_USD, H=BOJ_USD,
    # I=CB_Total, J=VIX, K=DXY, L=GLI_Flash, M=GLI_Full, N=CB_YoY, O=CB_3m_ann, P=ACWI
    # Q=Liquidity_Pct_Change, R=Trend_Signal

    # Apply conditional formatting ONLY to key actionable columns
    # Focus on what matters: the aggregate soup, not the ingredients
    # Key columns: CB_Total (I), GLI_Flash (L), GLI_Full (M), CB_YoY (N), CB_3m_ann (O), Liquidity_Pct_Change (Q)
    key_columns = {
        'I': 'CB_Total',           # Global liquidity total - the main signal
        'L': 'GLI_Flash',          # Flash index - short-term momentum
        'M': 'GLI_Full',           # Full index - macro regime
        'N': 'CB_YoY',             # Year-over-year growth
        'O': 'CB_3m_ann'           # 3-month annualized growth
    }

    for col_letter in key_columns.keys():
        cell_range = f"{col_letter}2:{col_letter}{max_row}"
        rule = ColorScaleRule(
            start_type="min", start_color="FF6B6B",  # Red for low values
            mid_type="percentile", mid_value=50, mid_color="FFD93D",  # Yellow for mid
            end_type="max", end_color="6BCF7F"  # Green for high values
        )
        worksheet.conditional_formatting.add(cell_range, rule)

    # Color-code Liquidity_Pct_Change column (Q) with directional meaning
    # Red = shrinking liquidity (tightening), Green = expanding (easing), Neutral for small moves
    liq_change_range = f"Q2:Q{max_row}"
    liq_change_rule = ColorScaleRule(
        start_type="num", start_value=-0.5, start_color="FF6B6B",  # Red for -0.5% or less (shrinking)
        mid_type="num", mid_value=0, mid_color="F5F5F5",  # Light gray for neutral/small moves
        end_type="num", end_value=0.5, end_color="6BCF7F"  # Green for +0.5% or more (expanding)
    )
    worksheet.conditional_formatting.add(liq_change_range, liq_change_rule)

    # Color-code the Trend_Signal column (last column)
    # Find the column letter for Trend_Signal
    trend_col_idx = len(weekly_export.columns) + 1  # +1 because index takes column A

    # Define fill colors
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Apply color to each cell in Trend_Signal column based on value
    for row in range(2, max_row + 1):  # Start from row 2 (skip header)
        cell = worksheet.cell(row=row, column=trend_col_idx)
        if cell.value == "Rising":
            cell.fill = green_fill
        elif cell.value == "Falling":
            cell.fill = red_fill
        elif cell.value == "Flat":
            cell.fill = yellow_fill

    # ----------------------------------------------------------------------
    # Add Excel Charts Sheet
    # ----------------------------------------------------------------------
    try:
        # Create a new worksheet for charts
        charts_ws = workbook.create_sheet("Charts")

        from openpyxl.chart import LineChart, Reference

        # Chart 1: Global Liquidity Index (CB_Total)
        chart1 = LineChart()
        chart1.title = "Global Liquidity Trend"
        chart1.y_axis.title = "CB Total (USD Billions)"
        chart1.x_axis.title = "Date"

        # Date categories (x-axis) - skip header row
        dates = Reference(worksheet, min_col=1, min_row=2, max_row=max_row)
        # Data values (y-axis) - CB_Total is column I (9)
        data1 = Reference(worksheet, min_col=9, min_row=1, max_row=max_row)

        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(dates)

        # Simplify axis configuration - use default category axis instead of DateAxis
        chart1.x_axis.number_format = 'd-mmm-yy'

        chart1.width = 20
        chart1.height = 10
        charts_ws.add_chart(chart1, "A1")

        # Chart 2: YoY Liquidity Growth
        chart2 = LineChart()
        chart2.title = "YoY Liquidity Growth (Above 0 = Easing, Below 0 = Tightening)"
        chart2.style = 2
        chart2.y_axis.title = "YoY Growth (%)"
        chart2.x_axis.title = "Date"

        # Data for CB_YoY (column N)
        data2 = Reference(worksheet, min_col=14, min_row=1, max_row=max_row, max_col=14)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(dates)

        # Simplify axis configuration
        chart2.x_axis.number_format = 'd-mmm-yy'

        chart2.legend = None  # Remove legend since it's just one series
        chart2.width = 20
        chart2.height = 10
        charts_ws.add_chart(chart2, "A25")

        # Chart 3: Short-Term Trend (% Change)
        chart3 = LineChart()
        chart3.title = "Week-over-Week Liquidity Change (Early Warning Signal)"
        chart3.style = 2
        chart3.y_axis.title = "Weekly % Change"
        chart3.x_axis.title = "Date"

        # Data for Liquidity_Pct_Change (column Q)
        data3 = Reference(worksheet, min_col=17, min_row=1, max_row=max_row, max_col=17)
        chart3.add_data(data3, titles_from_data=True)
        chart3.set_categories(dates)

        # Simplify axis configuration
        chart3.x_axis.number_format = 'd-mmm-yy'

        chart3.legend = None  # Remove legend since it's just one series
        chart3.width = 20
        chart3.height = 10
        charts_ws.add_chart(chart3, "A49")

        # Chart 4: Equities vs Liquidity Overlay (Dual Axis)
        # Use normalized overlay data from MSCI_Overlay sheet for proper scaling
        overlay_ws = writer.sheets["MSCI_Overlay"]
        overlay_max_row = overlay_ws.max_row

        chart4 = LineChart()
        chart4.title = "Global Liquidity vs Equities (Liquidity Leads, Stocks Follow)"
        chart4.style = 2
        chart4.y_axis.title = "Normalized Index (0-1)"
        chart4.x_axis.title = "Date"

        # Use overlay sheet dates and data (already normalized 0-1)
        overlay_dates = Reference(overlay_ws, min_col=1, min_row=2, max_row=overlay_max_row)

        # GLI data (column B in overlay sheet)
        gli_data = Reference(overlay_ws, min_col=2, min_row=1, max_row=overlay_max_row, max_col=2)
        chart4.add_data(gli_data, titles_from_data=True)
        chart4.set_categories(overlay_dates)

        # MSCI data (column C in overlay sheet)
        msci_data = Reference(overlay_ws, min_col=3, min_row=1, max_row=overlay_max_row, max_col=3)
        chart4.add_data(msci_data, titles_from_data=True)

        # Simplify axis configuration
        chart4.x_axis.number_format = 'd-mmm-yy'

        chart4.width = 20
        chart4.height = 10
        charts_ws.add_chart(chart4, "A73")

        print("Excel charts generated successfully")

    except Exception as e:
        print(f"Warning: Could not generate Excel charts: {e}")
        print("Excel file will be saved without embedded charts")

print("Done. Excel file saved with conditional formatting and charts:", EXCEL_OUTPUT)
